from openpyxl import load_workbook
import xlrd

# use package openpyxl
def read_by_openpyxl(filename: str):
    # load Excel
    try:
        wb = load_workbook(filename)
        sheet = wb.active
    except Exception as e:
        print(f"Error with {filename}: {e}")
        return
    for row in sheet.iter_rows(min_row=2, values_only=True):
        id,first_name,last_name, group_id = row
        print(f"First Name: {first_name}; Last Name: {last_name}; Group ID: {group_id}")

# use package xlrd, however it only support xls format
def read_by_xlrd(filename:str):
    # open Excel file
    try:
        workbook = xlrd.open_workbook(filename)
    except Exception as e:
        print(f"Error with {filename}: {e}")
        return
    # get worksheet, 0-the first worksheet
    sheet = workbook.sheet_by_index(0)
    # sheet.nrows - get the total number of rows
    for row in range(1,sheet.nrows):
        first_name = sheet.cell_value(row,1)
        last_name = sheet.cell_value(row,2)
        group_id = sheet.cell_value(row,3)
        print(f"First Name: {first_name}; Last Name: {last_name}; Group ID: {group_id}")


# read_by_openpyxl("../file/student.xlsx")
read_by_xlrd("../file/student2.xls")